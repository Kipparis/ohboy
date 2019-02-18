#!python

print("___________\nstarted\n")

import sys

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from openpyxl.styles.alignment import Alignment
from openpyxl.styles import PatternFill

from datetime import datetime, timedelta

def correct_table(ws):
    for col in ws.columns:
            max_length = 0
            column = col[0].column # Get the column name
            for cell in col:
                try: # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            # TODO: Тут выдаёт ошибку (
            # ws.column_dimensions[column].width = adjusted_width

days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
spec = ['name', 'cost', 'general cost']
colors = ['FFFFCC', 'CCFFCC', 'CCFFFF', 'CCCCFF', 'FFCCFF', 'FFCCCC', 'FF9999']


for arg in sys.argv:
    if arg in "-maketable":
        print("Making table")

        # создаём таблицу
        wb = Workbook()
        pays = wb.active
        pays.title = "Pays"

        pays["A1"] = "Days"
        pays["A2"] = "Date\\Type"


        # Merge cells and set names
        for i in range(1, 8):
            # Find two side columns to merge 
            curr_col = i * 3 - 1
            next_col = i * 3 + 1

            pays.merge_cells(start_row=1, end_row=1, start_column=curr_col, end_column=next_col)
            curr_cell = pays.cell(row=1, column=curr_col, value=days[i - 1])
            curr_cell.alignment = Alignment(horizontal='center')
            curr_cell.fill = PatternFill(bgColor=colors[i - 1], fill_type = "gray0625")

        # Give them spec
        for i in range(1, 8):
            for j in range(3):
                curr_cell = pays.cell(row=2, column=i * 3 - 1 + j, value=spec[j])
                curr_cell.alignment = Alignment(horizontal='center') 
                curr_cell.fill = PatternFill(bgColor=colors[i - 1], fill_type = "gray0625")
            

        correct_table(pays)

        wb["Pays"].freeze_panes = pays.cell(row=3, column=7 * 3 + 1)
        wb.save('fins.xlsx')
        print("All ok")

    # TODO: Add alignment and coloring
    if arg in "-add":
        wb = load_workbook('fins.xlsx')
        pays = wb['Pays']

        today = datetime.now()
        week_day = today.weekday()
        
        past_date = today - timedelta(days=week_day)
        future_date = today + timedelta(days=(7 - week_day))
        
        names = []
        costs = []
        print("type 'end' to add products to table")
        while True:
            name = input("Name your product:\n->")
            if name in "end":
                break

            cost = input("Type down its cost:\n->")
            if cost in "end":
                break

            names.append(name) 
            costs.append(cost)

        # Go for table and look table
        # check first column
        col = pays['A']

        check_row = len(col)

        if len(pays['A']) <= 2:
            # Don't fit    
            check_row = 3

            pays.cell(
                row=check_row, 
                column=1, 
                value="{}/{}/{}".format(past_date.day, past_date.month, past_date.year)
                )

        cell = pays.cell(row=check_row, column=1)
        while not cell.value:
            check_row -= 1
            cell = pays.cell(row=check_row, column=1)
        not_empty_row = check_row

        while '*' in str(cell.value):
            check_row -= 1
            cell = pays.cell(row=check_row, column=1)

        date_row = check_row
        cell_date = str(cell.value)


        cell_past_date = datetime(
            year=int(cell_date.split('/')[-1]),
            month=int(cell_date.split('/')[1]),
            day=int(cell_date.split('/')[0])                
        )

        date_row = 0

        if past_date.day == cell_past_date.day and past_date.month == cell_past_date.month and past_date.year == cell_past_date.year:
            start_row = check_row
            date_row = check_row
        else:
            # Have to move to another date
            start_row = not_empty_row + 1
            # Type there date
            pays.cell(
                row=start_row, 
                column=1, 
                value="{}/{}/{}".format(past_date.day, past_date.month, past_date.year)
                )
            date_row = start_row

        # Now start row pointer is one date cell
        # and we need to go to the column and find cell to start typing 
        # costs
        curr_col = 2 + week_day * 3

        # If not => move down and find empty than \
        # If that cell is empty => start typing product and type on left *

        start_row = len(pays['A']) + 1
        cell = pays.cell(row=start_row, column=curr_col)

        while not cell.value:
            start_row -= 1
            cell = pays.cell(row=start_row, column=curr_col)
        
        start_row += 1

        start_row = max(date_row, start_row)

        # Start typing name, cost and * on left column ( if she empty)
        for spend in zip(names, costs):
            pays.cell(row=start_row, column=curr_col, value=spend[0])
            pays.cell(row=start_row, column=curr_col+1, value=int(spend[1]))

            if not pays.cell(row=start_row, column=1).value:
                pays.cell(row=start_row, column=1, value='*')

            start_row += 1

        pays.merge_cells(start_row=date_row, end_row=start_row-1, start_column=curr_col+2, end_column=curr_col+2)
        curr_cell = pays.cell(row=date_row, column=curr_col+2, value="=SUM({0}{1}:{0}{2})".format(get_column_letter(curr_col+1), date_row, start_row-1))
        curr_cell.alignment = Alignment(vertical='center', horizontal='center') 

        correct_table(pays)

        wb.save('fins.xlsx')



print("\nended\n***********")