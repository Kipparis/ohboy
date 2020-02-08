#!python3

print("started")

from openpyxl import Workbook

wb = Workbook() # Creates workbook
ws = wb.active # Creates worksheet ( main one )

ws.title = "Pays" # Give him name

ws.sheet_properties.tabColor = "FFFFFF" # Set color of tab


# Writing sample data
ws['A4'] = 200
d = ws.cell(row=4, column=2, value="method")

# Accessing many sells
cell_range = ws['B4':'C5']
# print(cell_range)

for cols in cell_range:
    for cell in cols:
        cell.value = 4

# cell_range = 4 # Don't know, maybeee works

# Other accessing ws['C'], ws['C:D'], ws[10], ws[5:10]

# Iteration be like :p
# for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
# ...    for cell in row:
# ...        print(cell)
# or
# for row in ws.iter_cols...
wb.save('fins.xlsx')

print("ended")
